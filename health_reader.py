"""
Unified health data reader for DailyReport & WeeklyReport.

Reads HealthData.xlsx (Summarize sheet) via openpyxl, parses pipe-delimited
fields, validates previous-day status, and optionally syncs from Google Drive.
"""
from __future__ import annotations

import shutil
import subprocess
import sys
import tempfile
import time as _time
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any, Optional

import openpyxl


# =========================
# Excel schema
# =========================
SHEET_NAME = "Summarize"
DATA_START_ROW = 3  # row 2 has summary aggregates

COLUMN_MAP = {
    "date":             "data",
    "planned_ot":       "planned",
    "actual_ot":        "actual",
    "ot_diff":          "difference",
    "start_work":       "clock_in",
    "end_work":         "clock_out",
    "tw":               "TW",
    "workout_feedback": "WorkoutFeedback",
    "sleep_time":       "sleep_start+wake_time",
    "sleep_stages":     "InBed|Core|Deep|Rem|",
    "sleep_result":     "sleep_result",
    "nutrition_total":  "nutrition_total",
    "nutrition_delta":  "nutrition_delta",
    "weight":           "weight",
    "note":             "notes",
    "status":           "status",
}


# =========================
# Data structure
# =========================
@dataclass
class HealthRow:
    date: str
    planned_ot: str
    actual_ot: str
    ot_diff: str
    start_work: str
    end_work: str
    tw: str
    workout_feedback: str
    sleep_start: str
    wake_time: str
    inbed_min: str
    core_min: str
    deep_min: str
    rem_min: str
    sleep_actual_min: str
    sleep_debt_delta: str
    sleep_debt_total: str
    sleep_status: str
    kcal: str
    carb: str
    protein: str
    fat: str
    nutrition_delta_kcal: str
    nutrition_delta_carb: str
    nutrition_delta_protein: str
    nutrition_delta_fat: str
    weight: str
    note: str
    status: str


# =========================
# Basic utilities
# =========================
def safe_str(value: Any) -> str:
    return "" if value is None else str(value)


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\r", "").strip()


def single_line_text(value: Any) -> str:
    return " ".join(clean_text(value).split())


def normalize_date_cell(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = clean_text(value)
    if not text:
        return None
    text = (
        text.replace("年", "-").replace("月", "-").replace("日", "")
        .replace("/", "-").replace(".", "-").strip()
    )
    parts = text.split("-")
    if len(parts) != 3:
        return None
    try:
        return date(int(parts[0]), int(parts[1]), int(parts[2])).isoformat()
    except ValueError:
        return None


def format_scalar(value: Any) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, (int, float)):
        num = float(value)
        return str(int(num)) if num.is_integer() else str(num)
    return clean_text(value)


def format_clock(value: Any) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, datetime):
        return value.strftime("%H:%M")
    if isinstance(value, time):
        return value.strftime("%H:%M")
    if isinstance(value, timedelta):
        total_minutes = int(round(value.total_seconds() / 60))
        return f"{total_minutes // 60}:{total_minutes % 60:02d}"
    if isinstance(value, (int, float)):
        total_minutes = int(round(float(value) * 24 * 60))
        return f"{total_minutes // 60}:{total_minutes % 60:02d}"
    return clean_text(value)


def format_number(value: Any, digits: int) -> str:
    if value is None or value == "":
        return ""
    try:
        return f"{float(value):.{digits}f}"
    except (TypeError, ValueError):
        return clean_text(value)


def parse_nullable_float(text: str) -> Optional[float]:
    text = clean_text(text)
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def format_metric(value: Optional[float], digits: int, unit: str) -> str:
    if value is None:
        return "未记录"
    return f"{value:.{digits}f} {unit}"


def format_avg(values: list[float], digits: int, unit: str) -> str:
    if not values:
        return "未记录"
    avg = sum(values) / len(values)
    return f"{avg:.{digits}f} {unit}（{len(values)}天）"


# =========================
# Pipe-delimited parsing
# =========================
def parse_pipe_kv(text: Any) -> dict[str, str]:
    text = clean_text(text)
    if not text:
        return {}
    result = {}
    for part in text.split("|"):
        if "=" in part:
            k, _, v = part.partition("=")
            result[k.strip()] = v.strip()
    return result


def parse_bracket_float(text: str) -> str:
    text = clean_text(text)
    if not text:
        return ""
    if text.startswith("(") and text.endswith(")"):
        inner = text[1:-1]
        try:
            float(inner)
            return f"-{inner}"
        except ValueError:
            return text
    try:
        val = float(text)
        return f"+{text}" if val > 0 else text
    except ValueError:
        return text


def parse_nutrition_pipe(text: Any) -> tuple[str, str, str, str]:
    text = clean_text(text)
    if not text:
        return "", "", "", ""
    parts = text.split("|")

    def get(i: int) -> str:
        if i >= len(parts):
            return ""
        return parse_bracket_float(parts[i].strip())

    return get(0), get(1), get(2), get(3)


# =========================
# Date range
# =========================
def build_target_dates(report_date: date, lookback_days: int = 14) -> list[str]:
    return [(report_date - timedelta(days=i)).isoformat() for i in range(lookback_days, 0, -1)]


# =========================
# File copy with retry
# =========================
def try_copy_with_retry(src: Path, retries: int = 5, wait_seconds: int = 3) -> Path:
    last_error: Optional[Exception] = None
    for attempt in range(1, retries + 1):
        temp_path = None
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix=src.suffix) as tmp:
                temp_path = Path(tmp.name)
            shutil.copy2(src, temp_path)
            print(f"[Unity] 文件复制成功，第 {attempt} 次尝试: {temp_path}")
            return temp_path
        except PermissionError as e:
            last_error = e
            if temp_path and temp_path.exists():
                try:
                    temp_path.unlink()
                except Exception:
                    pass
            print(f"[Unity] 文件复制第 {attempt} 次失败")
            if attempt < retries:
                _time.sleep(wait_seconds)
        except Exception:
            if temp_path and temp_path.exists():
                try:
                    temp_path.unlink()
                except Exception:
                    pass
            raise
    if last_error:
        raise last_error
    raise RuntimeError("复制文件失败")


def try_close_workbook_in_excel(target_file: str | Path) -> bool:
    target_file = str(Path(target_file).resolve()).lower()
    try:
        import win32com.client  # type: ignore
    except ImportError:
        print("[Unity] 未安装 pywin32，跳过关闭工作簿")
        return False
    try:
        excel = win32com.client.GetObject(Class="Excel.Application")
    except Exception:
        print("[Unity] 当前没有可连接的 Excel 实例")
        return False
    try:
        workbooks = excel.Workbooks
        for i in range(workbooks.Count, 0, -1):
            wb = workbooks.Item(i)
            try:
                wb_path = str(Path(wb.FullName).resolve()).lower()
            except Exception:
                continue
            if wb_path == target_file:
                print(f"[Unity] 命中已打开工作簿: {wb.FullName}")
                wb.Close(SaveChanges=False)
                print("[Unity] 已关闭目标工作簿（未保存更改）")
                return True
    except Exception as e:
        print(f"[Unity] 尝试关闭工作簿失败: {e}")
    return False


def prepare_readable_copy(
    src: Path,
    retries: int = 5,
    wait_seconds: int = 3,
    allow_close_excel_workbook: bool = False,
) -> Path:
    try:
        return try_copy_with_retry(src, retries=retries, wait_seconds=wait_seconds)
    except PermissionError:
        print("[Unity] 多次重试后仍被占用")
    if allow_close_excel_workbook:
        print("[Unity] 尝试关闭 Excel 中的目标工作簿...")
        closed = try_close_workbook_in_excel(src)
        if closed:
            return try_copy_with_retry(src, retries=3, wait_seconds=2)
    raise PermissionError(f"文件仍无法读取: {src}")


# =========================
# Core reader (openpyxl)
# =========================
def load_health_rows(
    file_path: Path,
    report_date: date,
    lookback_days: int = 14,
    copy_retries: int = 5,
    copy_wait_seconds: int = 3,
    allow_close_excel_workbook: bool = False,
) -> tuple[list[str], list[HealthRow]]:
    file_path = Path(file_path)
    if not file_path.exists():
        raise FileNotFoundError(f"文件不存在: {file_path}")

    temp_copy = prepare_readable_copy(
        file_path,
        retries=copy_retries,
        wait_seconds=copy_wait_seconds,
        allow_close_excel_workbook=allow_close_excel_workbook,
    )
    try:
        import warnings
        warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
        keep_vba = file_path.suffix.lower() == ".xlsm"
        wb = openpyxl.load_workbook(temp_copy, data_only=True, keep_vba=keep_vba)
        if SHEET_NAME not in wb.sheetnames:
            raise ValueError(f"找不到工作表: {SHEET_NAME}")

        ws = wb[SHEET_NAME]
        target_dates = build_target_dates(report_date, lookback_days)
        target_set = set(target_dates)

        header = {clean_text(ws.cell(1, c).value): c for c in range(1, ws.max_column + 1)}

        def get_col(key: str) -> int:
            col_name = COLUMN_MAP[key]
            if col_name not in header:
                raise ValueError(f"Excel header 中找不到列: {col_name!r}（key={key}）")
            return header[col_name]

        col_date         = get_col("date")
        col_planned_ot   = get_col("planned_ot")
        col_actual_ot    = get_col("actual_ot")
        col_ot_diff      = get_col("ot_diff")
        col_start_work   = get_col("start_work")
        col_end_work     = get_col("end_work")
        col_tw           = get_col("tw")
        col_wf           = get_col("workout_feedback")
        col_sleep_time   = get_col("sleep_time")
        col_sleep_stages = get_col("sleep_stages")
        col_sleep_result = get_col("sleep_result")
        col_nut_total    = get_col("nutrition_total")
        col_nut_delta    = get_col("nutrition_delta")
        col_weight       = get_col("weight")
        col_note         = get_col("note")
        col_status       = get_col("status")

        rows: list[HealthRow] = []

        for row_idx in range(DATA_START_ROW, ws.max_row + 1):
            raw_date = ws.cell(row_idx, col_date).value
            normalized_date = normalize_date_cell(raw_date)
            if not normalized_date or normalized_date not in target_set:
                continue

            sleep_kv = parse_pipe_kv(ws.cell(row_idx, col_sleep_time).value)
            stages_kv = parse_pipe_kv(ws.cell(row_idx, col_sleep_stages).value)
            result_kv = parse_pipe_kv(ws.cell(row_idx, col_sleep_result).value)
            kcal, carb, protein, fat = parse_nutrition_pipe(ws.cell(row_idx, col_nut_total).value)
            d_kcal, d_carb, d_protein, d_fat = parse_nutrition_pipe(ws.cell(row_idx, col_nut_delta).value)

            rows.append(HealthRow(
                date=normalized_date,
                planned_ot=format_scalar(ws.cell(row_idx, col_planned_ot).value),
                actual_ot=format_scalar(ws.cell(row_idx, col_actual_ot).value),
                ot_diff=format_scalar(ws.cell(row_idx, col_ot_diff).value),
                start_work=format_clock(ws.cell(row_idx, col_start_work).value),
                end_work=format_clock(ws.cell(row_idx, col_end_work).value),
                tw=clean_text(ws.cell(row_idx, col_tw).value),
                workout_feedback=clean_text(ws.cell(row_idx, col_wf).value),
                sleep_start=sleep_kv.get("sleep", ""),
                wake_time=sleep_kv.get("wake", ""),
                inbed_min=stages_kv.get("inbed", ""),
                core_min=stages_kv.get("core", ""),
                deep_min=stages_kv.get("deep", ""),
                rem_min=stages_kv.get("rem", ""),
                sleep_actual_min=result_kv.get("actual", ""),
                sleep_debt_delta=result_kv.get("debt_delta", ""),
                sleep_debt_total=result_kv.get("debt_total", ""),
                sleep_status=result_kv.get("status", ""),
                kcal=kcal,
                carb=carb,
                protein=protein,
                fat=fat,
                nutrition_delta_kcal=d_kcal,
                nutrition_delta_carb=d_carb,
                nutrition_delta_protein=d_protein,
                nutrition_delta_fat=d_fat,
                weight=format_number(ws.cell(row_idx, col_weight).value, 1),
                note=clean_text(ws.cell(row_idx, col_note).value),
                status=clean_text(ws.cell(row_idx, col_status).value),
            ))

        rows.sort(key=lambda x: x.date)
        return target_dates, rows
    finally:
        temp_copy.unlink(missing_ok=True)


# =========================
# Validation
# =========================
def validate_previous_day_health_data(
    file_path: str | Path,
    report_date: date,
    copy_retries: int = 5,
    copy_wait_seconds: int = 3,
    allow_close_excel_workbook: bool = False,
) -> dict[str, Any]:
    target_date = (report_date - timedelta(days=1)).isoformat()

    _, rows = load_health_rows(
        file_path=Path(file_path),
        report_date=report_date,
        lookback_days=1,
        copy_retries=copy_retries,
        copy_wait_seconds=copy_wait_seconds,
        allow_close_excel_workbook=allow_close_excel_workbook,
    )

    row = next((r for r in rows if r.date == target_date), None)

    if row is None:
        return {
            "ok": False,
            "report_date": report_date.isoformat(),
            "target_date": target_date,
            "row_found": False,
            "status_value": "",
            "status_ok": False,
            "detail": f"前一天 {target_date} 没有找到任何健康记录行。",
        }

    status_value = single_line_text(row.status)
    status_ok = "已完成" in status_value

    return {
        "ok": status_ok,
        "report_date": report_date.isoformat(),
        "target_date": target_date,
        "row_found": True,
        "status_value": status_value,
        "status_ok": status_ok,
        "detail": (
            f"前一天 {target_date} 的状态列为已完成。"
            if status_ok
            else f"前一天 {target_date} 的状态列不是已完成，当前值: {status_value or '空白'}"
        ),
    }


def build_previous_day_validation_message(result: dict[str, Any]) -> str:
    status = "通过" if result.get("ok") else "失败"
    return "\n".join([
        "=== 前一天数据校验 ===",
        f"报表日期: {result.get('report_date', '')}",
        f"校验目标日期: {result.get('target_date', '')}",
        f"结果: {status}",
        f"详情: {result.get('detail', '')}",
    ])


# =========================
# Google Drive sync
# =========================
def sync_health_file_from_drive(gdrive_script: str | Path, gdrive_config: str | Path) -> None:
    script_path = Path(gdrive_script).resolve()
    config_path = Path(gdrive_config).resolve()

    if not script_path.exists():
        raise FileNotFoundError(f"找不到 Google Drive 同步脚本: {script_path}")
    if not config_path.exists():
        raise FileNotFoundError(f"找不到 Google Drive 配置文件: {config_path}")

    cmd = [sys.executable, str(script_path), str(config_path)]
    print(f"[Unity] Google Drive 同步: {' '.join(cmd)}")
    subprocess.run(cmd, check=True)


# =========================
# High-level entry point
# =========================
def load_health_data(
    file_path: str | Path,
    report_date: date,
    lookback_days: int = 14,
    copy_retries: int = 5,
    copy_wait_seconds: int = 3,
    allow_close_excel_workbook: bool = False,
    gdrive_config: str | Path | None = None,
    gdrive_script: str | Path | None = None,
) -> dict[str, Any]:
    """
    Unified entry point used by both daily and weekly pipelines.

    1. Validates previous-day status.
    2. Optionally syncs from Google Drive if validation fails.
    3. Reads health rows for the target date range.

    Returns dict with keys: target_dates, rows, validation_result.
    """
    file_path = Path(file_path)

    # Validate
    validation_result = validate_previous_day_health_data(
        file_path=file_path,
        report_date=report_date,
        copy_retries=copy_retries,
        copy_wait_seconds=copy_wait_seconds,
        allow_close_excel_workbook=allow_close_excel_workbook,
    )
    print(build_previous_day_validation_message(validation_result))
    print()

    # Sync if needed
    if not validation_result["ok"] and gdrive_config:
        script_path = (
            Path(gdrive_script).resolve()
            if gdrive_script
            else Path(__file__).with_name("gdrive_to_local_onedrive_v3.py")
        )
        print("=== 校验未通过，从 Google Drive 同步 ===")
        sync_health_file_from_drive(script_path, gdrive_config)
        print()

    # Read
    target_dates, rows = load_health_rows(
        file_path=file_path,
        report_date=report_date,
        lookback_days=lookback_days,
        copy_retries=copy_retries,
        copy_wait_seconds=copy_wait_seconds,
        allow_close_excel_workbook=allow_close_excel_workbook,
    )

    return {
        "target_dates": target_dates,
        "rows": rows,
        "validation_result": validation_result,
    }
